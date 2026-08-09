"""XLSX inspection via openpyxl. Spreadsheets have sheets, not pages - no
`page_count` here (Excel's print-page-breaks are fragile/version-dependent
and not worth the false confidence of a number)."""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from dokuma.types import DocumentInfo

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


def _dimensions_str(sheet: Worksheet) -> str:
    # read_only worksheets have no `.dimensions` (that's computed from the
    # full parsed sheet, which read_only mode deliberately avoids) - build
    # the same "A1:D20" shape from the min/max row/col it does expose.
    from openpyxl.utils import get_column_letter

    start = f"{get_column_letter(sheet.min_column or 1)}{sheet.min_row or 1}"
    end = f"{get_column_letter(sheet.max_column or 1)}{sheet.max_row or 1}"
    return f"{start}:{end}"


class XlsxInspector:
    format = "xlsx"

    def inspect(self, path: Path) -> DocumentInfo:
        import openpyxl

        workbook = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        try:
            sheet_names = list(workbook.sheetnames)
            sheet_dimensions = {name: _dimensions_str(workbook[name]) for name in sheet_names}

            props = workbook.properties
            metadata = {
                k: str(v)
                for k, v in {
                    "creator": props.creator,
                    "subject": props.subject,
                    "created": props.created,
                    "modified": props.modified,
                }.items()
                if v
            }

            return DocumentInfo(
                path=path,
                format="xlsx",
                size_bytes=path.stat().st_size,
                sheet_count=len(sheet_names),
                sheet_names=sheet_names,
                sheet_dimensions=sheet_dimensions,
                title=props.title or None,
                metadata=metadata,
            )
        finally:
            workbook.close()
