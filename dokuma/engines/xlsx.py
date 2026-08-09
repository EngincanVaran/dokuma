"""XLSX extraction via openpyxl. Each sheet becomes one "table" region -
the natural structure of a spreadsheet, unlike PDF/DOCX where a table is
one kind of content among several.

`page` is repurposed as the 1-indexed sheet position, not a print page -
XLSX doesn't have print pages in any reliable sense (same reasoning
inspectors/xlsx.py uses for not reporting `page_count`). `bbox` stays
`None` - there's no meaningful pixel-position geometry for spreadsheet
cells. Empty sheets are skipped: verified directly that a sheet with no
data returns an empty `iter_rows()` in read-only mode, so "has any rows at
all" is a sufficient, cheap check.

Real bug, found via real-document testing: **openpyxl always returns
`datetime.datetime` for a date-backed cell, even one written as a plain
`datetime.date`** - confirmed directly, `sheet["A1"] = datetime.date(...)`
reads back as `datetime.datetime(..., 0, 0)`. Rendered with plain `str()`,
that's a spurious `00:00:00` on every date cell, not just genuine
midnight timestamps. Fixed with a heuristic, not a guarantee: a
`datetime.datetime` whose time component is exactly midnight renders as
date-only; a real midnight timestamp (rare in spreadsheet data, common
case is date-only cells) would render the same way and lose its time -
an accepted, documented tradeoff, not something openpyxl gives us a way
to distinguish.

Image regions need a **second, non-read-only workbook handle** - verified
directly that `read_only=True` worksheets don't expose `_images` at all
(`ReadOnlyWorksheet` never parses the drawing XML; `AttributeError` on
access), so the fast read-only path used for cell data can't also give us
embedded pictures. Opening a second full workbook costs more memory for
very large files - an accepted tradeoff for now, not a free win. Images
come back as their original embedded bytes (not rendered, unlike
PdfPlumberEngine), page = sheet index and bbox = None like table regions
(no meaningful pixel geometry for a spreadsheet), ordered after that
sheet's table region, sorted by anchor cell (row then column) - the
anchor is the closest thing to a "position" a spreadsheet image has, not
real reading order. `worksheet._images` / `image._data()` are private,
undocumented openpyxl API - the only way to read back an existing
workbook's pictures (`add_image()` and openpyxl's public docs are
write-oriented) - same class of tradeoff `DocxEngine` accepts reaching
into `paragraph._element` for the same reason.

`ExtractConfig(extract_images=False)` skips opening the second workbook
handle entirely - the real, avoidable cost this knob exists for, not just
the resulting regions.
"""

from __future__ import annotations

import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

from dokuma.config import ExtractConfig
from dokuma.engines.base import Engine
from dokuma.types import ExtractionResult, Region

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime.datetime) and value.time() == datetime.time.min:
        return value.date().isoformat()
    return str(value)


def _rows_to_html(rows: list[Any]) -> str:
    rows_html = []
    for row in rows:
        cells = "".join(f"<td>{_cell_str(cell.value)}</td>" for cell in row)
        rows_html.append(f"<tr>{cells}</tr>")
    return "<table>" + "".join(rows_html) + "</table>"


def _sheet_images(sheet: Worksheet) -> list[Any]:
    images: list[Any] = getattr(sheet, "_images", [])
    return sorted(images, key=lambda img: (img.anchor._from.row, img.anchor._from.col))


class XlsxEngine(Engine):
    name = "openpyxl"
    format = "xlsx"

    def _extract(self, path: Path, config: ExtractConfig) -> ExtractionResult:
        import openpyxl

        workbook = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        workbook_with_images = (
            openpyxl.load_workbook(str(path), data_only=True) if config.extract_images else None
        )
        try:
            regions: list[Region] = []
            for index, name in enumerate(workbook.sheetnames):
                sheet = workbook[name]
                rows = list(sheet.iter_rows())
                images = _sheet_images(workbook_with_images[name]) if workbook_with_images else []
                if not rows and not images:
                    continue

                if rows:
                    regions.append(
                        Region(
                            category="table",
                            content=_rows_to_html(rows),
                            page=index + 1,
                            order=len(regions),
                        )
                    )
                for image in images:
                    regions.append(
                        Region(
                            category="image",
                            content="",
                            page=index + 1,
                            order=len(regions),
                            image_bytes=image._data(),
                            image_format=image.format,
                        )
                    )
        finally:
            workbook.close()
            if workbook_with_images is not None:
                workbook_with_images.close()

        return ExtractionResult(
            path=path,
            format=self.format,
            engine=self.name,
            regions=regions,
        )
