from __future__ import annotations

from pathlib import Path

from dokuma.engines.xlsx import XlsxEngine


def test_xlsx_engine_one_table_region_per_sheet(sample_xlsx: Path) -> None:
    # sample_xlsx (conftest.py) has two sheets with data: "Data", "Notes".
    result = XlsxEngine().extract(sample_xlsx)

    assert result.error is None
    assert result.engine == "openpyxl"
    assert [r.category for r in result.regions] == ["table", "table"]
    assert [r.page for r in result.regions] == [1, 2]
    assert [r.order for r in result.regions] == [0, 1]


def test_xlsx_engine_table_content(sample_xlsx: Path) -> None:
    result = XlsxEngine().extract(sample_xlsx)

    data_table, notes_table = result.tables
    assert "Name" in data_table
    assert "widgets" in data_table
    assert "42" in data_table
    assert "just a note" in notes_table


def test_xlsx_engine_no_bbox(sample_xlsx: Path) -> None:
    result = XlsxEngine().extract(sample_xlsx)
    assert all(r.bbox is None for r in result.regions)


def test_xlsx_engine_skips_empty_sheets(tmp_path: Path) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    workbook.active.title = "HasData"
    workbook.active["A1"] = "content"
    workbook.create_sheet("Empty")
    path = tmp_path / "with_empty_sheet.xlsx"
    workbook.save(str(path))

    result = XlsxEngine().extract(path)

    assert len(result.regions) == 1
    assert "content" in result.tables[0]


def test_xlsx_engine_rejects_non_xlsx_file(dense_text_pdf: Path) -> None:
    result = XlsxEngine().extract(dense_text_pdf)

    assert result.error is not None
    assert "xlsx" in result.error
    assert "pdf" in result.error


def test_xlsx_engine_renders_gap_cell_as_empty_not_literal_none(tmp_path: Path) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "x"
    workbook.active["C1"] = "y"  # B1 stays empty - a real gap cell, not just an unset row
    path = tmp_path / "gap.xlsx"
    workbook.save(str(path))

    result = XlsxEngine().extract(path)

    assert result.tables[0] == "<table><tr><td>x</td><td></td><td>y</td></tr></table>"


def test_xlsx_engine_date_only_cell_has_no_spurious_time(tmp_path: Path) -> None:
    # Real bug, fixed: openpyxl always returns datetime.datetime, even for
    # a cell written as datetime.date - rendering it with plain str() put
    # a spurious "00:00:00" on every date cell, found via real-document
    # testing.
    import datetime

    import openpyxl

    workbook = openpyxl.Workbook()
    workbook.active["A1"] = datetime.date(2026, 1, 12)
    path = tmp_path / "dates.xlsx"
    workbook.save(str(path))

    result = XlsxEngine().extract(path)

    assert "2026-01-12" in result.tables[0]
    assert "00:00:00" not in result.tables[0]


def test_xlsx_engine_real_timestamp_keeps_its_time(tmp_path: Path) -> None:
    # A cell with a genuine non-midnight time is left alone - only the
    # midnight-heuristic case gets the date-only treatment.
    import datetime

    import openpyxl

    workbook = openpyxl.Workbook()
    workbook.active["A1"] = datetime.datetime(2026, 1, 12, 14, 30, 0)
    path = tmp_path / "timestamps.xlsx"
    workbook.save(str(path))

    result = XlsxEngine().extract(path)

    assert "14:30:00" in result.tables[0]


def test_xlsx_engine_extracts_image_after_sheets_table_region(xlsx_with_image: Path) -> None:
    # xlsx_with_image (conftest.py): sheet "Data" has cells + one image,
    # sheet "JustAPicture" has ONLY an image, no cell data at all.
    result = XlsxEngine().extract(xlsx_with_image)

    assert result.error is None
    categories_by_page = [(r.category, r.page) for r in result.regions]
    assert categories_by_page == [
        ("table", 1),
        ("image", 1),
        ("image", 2),
    ]
    assert [r.order for r in result.regions] == [0, 1, 2]


def test_xlsx_engine_image_only_sheet_is_not_skipped(xlsx_with_image: Path) -> None:
    # Real bug, fixed: the old skip check was `if not rows: continue`,
    # which would have silently dropped a sheet with a picture but no
    # cell data - found via real-document testing (openpyxl embedding an
    # image doesn't show up as a "row" at all).
    result = XlsxEngine().extract(xlsx_with_image)
    picture_only_regions = [r for r in result.regions if r.page == 2]

    assert len(picture_only_regions) == 1
    assert picture_only_regions[0].category == "image"


def test_xlsx_engine_image_bytes_are_the_real_original_file(xlsx_with_image: Path) -> None:
    import io

    from PIL import Image

    result = XlsxEngine().extract(xlsx_with_image)
    image_region = next(r for r in result.regions if r.category == "image")

    assert image_region.image_format == "png"
    decoded = Image.open(io.BytesIO(image_region.image_bytes))
    assert decoded.format == "PNG"
    assert decoded.size == (60, 40)


def test_xlsx_engine_no_bbox_for_image_regions(xlsx_with_image: Path) -> None:
    result = XlsxEngine().extract(xlsx_with_image)
    assert all(r.bbox is None for r in result.regions)
